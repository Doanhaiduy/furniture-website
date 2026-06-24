import openpyxl
import os
import sys
from copy import copy
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

SCRATCH_DIR = "C:/Users/DELL/.gemini/antigravity/brain/9f6e44d4-94f3-4e29-a2d0-d4e30f923a9d/scratch"
os.makedirs(SCRATCH_DIR, exist_ok=True)
SCREENSHOT_DIR = os.path.join("artifacts", "screenshots", "pages", "desktop")

COL_WIDTHS = {'A': 8, 'B': 20, 'C': 20, 'D': 22, 'E': 32, 'F': 32, 'G': 68}

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def copy_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = copy(src_cell.number_format)
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

def apply_borders_and_wrap(sheet, start_row, end_row):
    for r in range(start_row, end_row + 1):
        for col in range(1, 8):
            cell = sheet.cell(row=r, column=col)
            cell.border = thin_border
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=True, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

def adjust_row_heights(sheet, start_row, end_row):
    for r in range(start_row, end_row + 1):
        max_lines = 1
        for col in range(1, 8):
            cell_val = sheet.cell(row=r, column=col).value
            if cell_val and isinstance(cell_val, str):
                lines = len(cell_val.split('\n'))
                if lines > max_lines:
                    max_lines = lines
        sheet.row_dimensions[r].height = max(20, max_lines * 16 + 10)

def unify_font(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=7):
        for cell in row:
            if cell.font:
                cell.font = Font(name='Segoe UI', size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
            else:
                cell.font = Font(name='Segoe UI', size=11)

def find_screenshot_placeholder(sheet):
    for r in range(1, 40):
        for c in range(1, 8):
            cell = sheet.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and "[CHÈN ẢNH" in cell.value:
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        return min_row, min_col
                return r, c
    return 12, 1

def insert_resized_screenshot(sheet, screenshot_filename, screen_code):
    if not screenshot_filename:
        print(f"[{screen_code}] No screenshot filename provided.")
        return 0
    row_idx, col_idx = find_screenshot_placeholder(sheet)
    img_path = os.path.join(SCREENSHOT_DIR, screenshot_filename)
    if os.path.exists(img_path):
        try:
            pil_img = PILImage.open(img_path)
            orig_w, orig_h = pil_img.size
            new_w = 1000
            new_h = int((new_w / orig_w) * orig_h)
            resized_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
            
            # Slicing logic
            slice_h = 800
            slices = []
            if new_h <= 1200:
                slices.append(resized_img)
            else:
                num_slices = (new_h + slice_h - 1) // slice_h
                # Cap at 4 slices to avoid giant excel sheet
                num_slices = min(num_slices, 4)
                for i in range(num_slices):
                    box_y1 = i * slice_h
                    box_y2 = min((i + 1) * slice_h, new_h)
                    slice_img = resized_img.crop((0, box_y1, new_w, box_y2))
                    slices.append(slice_img)
            
            # Unmerge placeholder cell
            placeholder_coord = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
            for merged_range in list(sheet.merged_cells.ranges):
                if placeholder_coord in merged_range:
                    sheet.unmerge_cells(str(merged_range))
            
            # Clear text in the placeholder area (usually rows 9-28)
            for r in range(9, 29):
                for c in range(1, 8):
                    sheet.cell(row=r, column=c).value = None
            
            rows_per_slice = 30
            total_rows_to_insert = len(slices) * rows_per_slice - 1
            
            # Insert rows to make space
            if total_rows_to_insert > 0:
                sheet.insert_rows(row_idx + 1, total_rows_to_insert)
            
            # Reset row heights and clear values
            for r in range(row_idx, row_idx + len(slices) * rows_per_slice):
                sheet.row_dimensions[r].height = 20
                for c in range(1, 8):
                    sheet.cell(row=r, column=c).value = None
            
            # Save slices and insert
            for idx, slice_img in enumerate(slices):
                temp_path = os.path.join(SCRATCH_DIR, f"temp_{screen_code}_part{idx+1}.png")
                slice_img.save(temp_path)
                img = OpenpyxlImage(temp_path)
                
                target_row = row_idx + idx * rows_per_slice
                cell_coord = openpyxl.utils.get_column_letter(col_idx) + str(target_row)
                sheet.add_image(img, cell_coord)
                print(f"  [OK] Slice {idx+1}: {screenshot_filename} -> {cell_coord}")
                
            return total_rows_to_insert
        except Exception as e:
            print(f"  [ERR] Screenshot insert error: {e}")
            return 0
    else:
        print(f"  [WARN] Screenshot NOT found: {img_path}")
        return 0

def populate_sheet(wb, data, screenshot_filename):
    screen_code = data['code']
    sheet_name = f"{screen_code}_{data['slug_name']}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    print(f"[{screen_code}] Creating {sheet_name}")
    sheet = wb.copy_worksheet(wb['Template_Man_Hinh'])
    sheet.title = sheet_name
    sheet.sheet_view.showGridLines = True
    try:
        sheet.views.sheetView[0].showGridLines = True
    except Exception:
        pass
    sheet.cell(row=1, column=1).value = f"ĐẶC TẢ CHI TIẾT MÀN HÌNH: {data['name'].upper()}"
    sheet.cell(row=4, column=2).value = screen_code
    sheet.cell(row=4, column=5).value = data['url']
    sheet.cell(row=5, column=2).value = data['system_type']
    sheet.cell(row=5, column=5).value = "AI Agent & Developer"
    sheet.cell(row=6, column=2).value = "Đã gen"
    sheet.cell(row=6, column=5).value = data['date']
    
    # Slicing and dynamic row insertion offset
    offset = insert_resized_screenshot(sheet, screenshot_filename, screen_code)
    
    ui_elements = data['ui_elements']
    num_ui = len(ui_elements)
    ui_elements_start_row = 30 + offset
    
    if num_ui > 1:
        sheet.insert_rows(ui_elements_start_row + 1, num_ui - 1)
        for offset_ui in range(1, num_ui):
            r = ui_elements_start_row + offset_ui
            for col in range(1, 8):
                copy_style(sheet.cell(row=ui_elements_start_row, column=col), sheet.cell(row=r, column=col))
            sheet.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            sheet.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            
    for idx, elem in enumerate(ui_elements):
        r = ui_elements_start_row + idx
        sheet.cell(row=r, column=1).value = elem[0]
        sheet.cell(row=r, column=2).value = elem[1]
        sheet.cell(row=r, column=4).value = elem[2]
        sheet.cell(row=r, column=5).value = elem[3]
        sheet.cell(row=r, column=7).value = elem[4]
        
    apply_borders_and_wrap(sheet, ui_elements_start_row, ui_elements_start_row + num_ui - 1)
    adjust_row_heights(sheet, ui_elements_start_row, ui_elements_start_row + num_ui - 1)
    
    workflow_template_row = ui_elements_start_row + 4 + (num_ui - 1)
    workflows = data['workflows']
    num_wf = len(workflows)
    
    if num_wf > 1:
        sheet.insert_rows(workflow_template_row + 1, num_wf - 1)
        for offset_wf in range(1, num_wf):
            r = workflow_template_row + offset_wf
            for col in range(1, 8):
                copy_style(sheet.cell(row=workflow_template_row, column=col), sheet.cell(row=r, column=col))
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            sheet.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
            
    for idx, wf in enumerate(workflows):
        r = workflow_template_row + idx
        sheet.cell(row=r, column=1).value = wf[0]
        sheet.cell(row=r, column=4).value = wf[1]
        
    apply_borders_and_wrap(sheet, workflow_template_row, workflow_template_row + num_wf - 1)
    adjust_row_heights(sheet, workflow_template_row, workflow_template_row + num_wf - 1)
    
    user_guide_row = workflow_template_row + 3 + (num_wf - 1)
    sheet.cell(row=user_guide_row, column=1).value = data['user_guide']
    apply_borders_and_wrap(sheet, user_guide_row, user_guide_row + 3)
    adjust_row_heights(sheet, user_guide_row, user_guide_row + 3)
    
    for col_letter, width in COL_WIDTHS.items():
        sheet.column_dimensions[col_letter].width = width
    unify_font(sheet)
    print(f"  [DONE] {sheet_name}\n")

def update_index_sheet(sheet, all_screen_entries, date_str):
    # Clear any old entries starting from row 15
    for r in range(15, 60):
        for col in range(1, 8):
            sheet.cell(row=r, column=col).value = None
            
    data_start_row = 15
    done_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    done_font = Font(name='Segoe UI', size=10, color="2E7D32")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    for i, entry in enumerate(all_screen_entries):
        r = data_start_row + i
        sheet.cell(row=r, column=1).value = entry['stt']
        sheet.cell(row=r, column=2).value = entry['code']
        sheet.cell(row=r, column=3).value = entry['name']
        sheet.cell(row=r, column=4).value = entry['url']
        sheet.cell(row=r, column=5).value = entry['component']
        sheet.cell(row=r, column=6).value = "Đã gen"
        sheet.cell(row=r, column=7).value = f"Hoàn thành tự động bởi AI Agent ({date_str})"
        for col in range(1, 8):
            cell = sheet.cell(row=r, column=col)
            cell.fill = done_fill
            cell.font = done_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.sheet_view.showGridLines = True
    try:
        sheet.views.sheetView[0].showGridLines = True
    except Exception:
        pass
    print(f"  [OK] Index sheet updated with {len(all_screen_entries)} entries.")

def main():
    print("=== Start Generating SRS Excel Documents ===")
    
    # Import srs data module
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from srs_data import client_screens_data, admin_screens_data, client_index_entries, admin_index_entries

    client_template_path = os.path.join("artifacts", "client_srs_template.xlsx")
    admin_template_path = os.path.join("artifacts", "admin_srs_template.xlsx")

    # --------------------------------------------
    # 1. GENERATE CLIENT SRS
    # --------------------------------------------
    if os.path.exists(client_template_path):
        print(f"\nLoading {client_template_path}...")
        wb_client = openpyxl.load_workbook(client_template_path)

        for data in client_screens_data:
            populate_sheet(wb_client, data, data.get('screenshot'))

        if "Thông tin chung" in wb_client.sheetnames:
            update_index_sheet(wb_client["Thông tin chung"], client_index_entries, "2026-06-22")

        wb_client.save(client_template_path)
        print(f"\n[SAVED] {client_template_path}")
    else:
        print(f"[ERROR] {client_template_path} not found.")

    # --------------------------------------------
    # 2. GENERATE ADMIN SRS
    # --------------------------------------------
    if os.path.exists(admin_template_path):
        print(f"\nLoading {admin_template_path}...")
        wb_admin = openpyxl.load_workbook(admin_template_path)

        for data in admin_screens_data:
            populate_sheet(wb_admin, data, data.get('screenshot'))

        if "Thông tin chung" in wb_admin.sheetnames:
            update_index_sheet(wb_admin["Thông tin chung"], admin_index_entries, "2026-06-22")

        wb_admin.save(admin_template_path)
        print(f"\n[SAVED] {admin_template_path}")
    else:
        print(f"[ERROR] {admin_template_path} not found.")

    print("\n=== All SRS Excel Documents Generated Successfully! ===")
    print(f"Client SRS: {client_template_path}")
    print(f"Admin SRS:  {admin_template_path}")

if __name__ == "__main__":
    main()
