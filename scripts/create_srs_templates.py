import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def setup_sheet_style(ws):
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

def create_title_section(ws, sheet_name, url_path, screen_id, is_admin=False):
    # Font definitions
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=11, bold=True, color='1E293B')
    label_font = Font(name='Segoe UI', size=10, bold=True, color='475569')
    val_font = Font(name='Segoe UI', size=10, color='0F172A')
    
    # Fills
    title_fill = PatternFill(start_color='1E3A8A' if not is_admin else '065F46', end_color='1E3A8A' if not is_admin else '065F46', fill_type='solid')
    info_header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Main Title
    ws.merge_cells('A1:G2')
    title_cell = ws['A1']
    title_cell.value = f"ĐẶC TẢ CHI TIẾT MÀN HÌNH: {sheet_name.upper()}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply fill and border to merged cell range
    for row in range(1, 3):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = title_fill
            cell.border = thin_border

    # 2. General Information Section
    info_items = [
        ("Mã màn hình", screen_id, "Đường dẫn URL", url_path),
        ("Loại hệ thống", "Client Website" if not is_admin else "Admin CMS Portal", "Người thực hiện", "AI Agent & Developer"),
        ("Trạng thái", "Bản thảo (Draft)", "Ngày cập nhật", "2026-06-22")
    ]
    
    current_row = 4
    for label1, val1, label2, val2 in info_items:
        # Col A - B
        ws.cell(row=current_row, column=1, value=label1).font = label_font
        ws.cell(row=current_row, column=1).fill = info_header_fill
        ws.cell(row=current_row, column=1).border = thin_border
        
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        val1_cell = ws.cell(row=current_row, column=2, value=val1)
        val1_cell.font = val_font
        val1_cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in range(2, 4):
            ws.cell(row=current_row, column=col).border = thin_border
            
        # Col D - E
        ws.cell(row=current_row, column=4, value=label2).font = label_font
        ws.cell(row=current_row, column=4).fill = info_header_fill
        ws.cell(row=current_row, column=4).border = thin_border
        
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
        val2_cell = ws.cell(row=current_row, column=5, value=val2)
        val2_cell.font = val_font
        val2_cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in range(5, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
    return current_row + 1

def create_section_header(ws, row, title, is_admin=False):
    section_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
    fill_color = '2563EB' if not is_admin else '0D9488'
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = fill
    cell.alignment = Alignment(vertical='center', indent=1)
    
    # Border for the merged row
    thin_border = Border(bottom=Side(style='medium', color='1E3A8A' if not is_admin else '0F766E'))
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = thin_border
        
    ws.row_dimensions[row].height = 26
    return row + 1

def create_ui_spec_table(ws, start_row, items, is_admin=False):
    # Headers
    headers = [
        "STT", 
        "Thành phần Giao diện (UI Element)", 
        "Loại Control", 
        "Mô tả Thiết kế & Định dạng (Design & Validation)", 
        "Nhiệp vụ & Xử lý sự kiện (Business Logic & Actions)"
    ]
    
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    header_fill_color = '3B82F6' if not is_admin else '14B8A6'
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Write headers
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=1) # STT
    ws.cell(row=start_row, column=1, value=headers[0]).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3) # UI Element
    ws.cell(row=start_row, column=2, value=headers[1]).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    ws.cell(row=start_row, column=4, value=headers[2]).alignment = Alignment(horizontal='center', vertical='center') # Control Type
    
    ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=6) # Design & Validation
    ws.cell(row=start_row, column=5, value=headers[3]).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.cell(row=start_row, column=7, value=headers[4]).alignment = Alignment(horizontal='left', vertical='center') # Business Logic & Actions
    
    # Format headers
    for col in range(1, 8):
        cell = ws.cell(row=start_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    ws.row_dimensions[start_row].height = 28
    
    # Write data
    text_font = Font(name='Segoe UI', size=10, color='334155')
    bold_text_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    current_row = start_row + 1
    
    for idx, (element, ctrl_type, design, business) in enumerate(items, 1):
        # STT
        stt_cell = ws.cell(row=current_row, column=1, value=idx)
        stt_cell.font = bold_text_font
        stt_cell.alignment = Alignment(horizontal='center', vertical='top')
        stt_cell.border = thin_border
        
        # UI Element (merged Col B & C)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        el_cell = ws.cell(row=current_row, column=2, value=element)
        el_cell.font = bold_text_font
        el_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        
        # Control Type (Col D)
        ctrl_cell = ws.cell(row=current_row, column=4, value=ctrl_type)
        ctrl_cell.font = text_font
        ctrl_cell.alignment = Alignment(horizontal='center', vertical='top')
        ctrl_cell.border = thin_border
        
        # Design & Validation (merged Col E & F)
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        ds_cell = ws.cell(row=current_row, column=5, value=design)
        ds_cell.font = text_font
        ds_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Business Logic & Actions (Col G)
        bs_cell = ws.cell(row=current_row, column=7, value=business)
        bs_cell.font = text_font
        bs_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set border for all merged cells
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            
        # Estimate height based on text length
        max_len = max(len(element), len(design), len(business))
        row_height = max(35, min(120, max_len // 2))
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
        
    return current_row + 1

def create_business_flow_section(ws, start_row, flows, is_admin=False):
    header_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    text_font = Font(name='Segoe UI', size=10, color='334155')
    bold_text_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Table Header
    ws.cell(row=start_row, column=1, value="Luồng nghiệp vụ / Use case").font = header_font
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    
    ws.cell(row=start_row, column=4, value="Các bước thực hiện (Step-by-step)").font = header_font
    ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=7)
    
    fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for col in range(1, 8):
        ws.cell(row=start_row, column=col).fill = fill
        ws.cell(row=start_row, column=col).border = thin_border
        
    ws.row_dimensions[start_row].height = 24
    
    current_row = start_row + 1
    for flow_name, steps in flows:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        f_cell = ws.cell(row=current_row, column=1, value=flow_name)
        f_cell.font = bold_text_font
        f_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=7)
        s_cell = ws.cell(row=current_row, column=4, value=steps)
        s_cell.font = text_font
        s_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = thin_border
            
        max_len = max(len(flow_name), len(steps))
        row_height = max(35, min(150, max_len // 2))
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
        
    return current_row + 1

def create_user_guide_section(ws, start_row, steps):
    text_font = Font(name='Segoe UI', size=10, color='334155')
    bold_text_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+3, end_column=7)
    cell = ws.cell(row=start_row, column=1, value=steps)
    cell.font = text_font
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    for r in range(start_row, start_row + 4):
        ws.row_dimensions[r].height = 20
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border
            
    return start_row + 5

def create_image_placeholder_section(ws, start_row, screenshot_note):
    label_font = Font(name='Segoe UI', size=10, bold=True, color='475569')
    val_font = Font(name='Segoe UI', size=10, italic=True, color='64748B')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws.cell(row=start_row, column=1, value="Ảnh chụp màn hình (UI Screenshot)").font = label_font
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    ws.cell(row=start_row, column=1).border = thin_border
    ws.row_dimensions[start_row].height = 24
    
    # Place holder area for image (rows start_row+1 to start_row+15)
    ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+15, end_column=7)
    img_cell = ws.cell(row=start_row+1, column=1, value=f"[CHÈN ẢNH CHỤP MÀN HÌNH TẠI ĐÂY]\n\nTên file ảnh mẫu:\n{screenshot_note}\n\n(AI Agent sẽ tự động chụp ảnh thực tế của web đang chạy và chèn đè lên vùng này)")
    img_cell.font = val_font
    img_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    img_cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
    
    for r in range(start_row+1, start_row+16):
        ws.row_dimensions[r].height = 20
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border
            
    return start_row + 17

def create_info_intro_sheet(wb, is_admin=False):
    ws = wb.active
    ws.title = "Thông tin chung"
    setup_sheet_style(ws)
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=12, bold=True, color='1E293B')
    text_font = Font(name='Segoe UI', size=10, color='334155')
    bold_text_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
    
    title_fill = PatternFill(start_color='1E3A8A' if not is_admin else '065F46', end_color='1E3A8A' if not is_admin else '065F46', fill_type='solid')
    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws.merge_cells('A1:G2')
    title_cell = ws['A1']
    title_cell.value = f"TÀI LIỆU ĐẶC TẢ THIẾT KẾ CHI TIẾT MÀN HÌNH - PHÂN HỆ {'CLIENT' if not is_admin else 'ADMIN'}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for r in range(1, 3):
        ws.row_dimensions[r].height = 20
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = title_fill
            ws.cell(row=r, column=c).border = thin_border
            
    # Intro info
    ws.cell(row=4, column=1, value="GIỚI THIỆU TÀI LIỆU").font = section_font
    ws.merge_cells('A4:G4')
    ws.row_dimensions[4].height = 24
    
    intro_text = (
        "Tài liệu này chứa các sheet đặc tả chi tiết cho từng màn hình của phân hệ "
        f"{'Client (dành cho khách hàng mua sắm đồ gỗ và thiết bị vệ sinh)' if not is_admin else 'Admin Portal (dành cho quản trị viên vận hành cửa hàng)'}.\n"
        "Mỗi sheet tương ứng với một màn hình thực tế trên ứng dụng web Next.js đang chạy.\n\n"
        "Nội dung đặc tả bao gồm:\n"
        "1. Thông tin chung (Mã màn hình, URL, người thực hiện).\n"
        "2. Ảnh chụp màn hình giao diện thực tế (UI Screenshot).\n"
        "3. Danh sách đặc tả từng UI Element (Loại control, validation, logic thiết kế).\n"
        "4. Mô tả nghiệp vụ & Luồng xử lý chi tiết (Business Logic & Workflows).\n"
        "5. Hướng dẫn sử dụng chi tiết dành cho người dùng."
    )
    ws.cell(row=5, column=1, value=intro_text).font = text_font
    ws.cell(row=5, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells('A5:G10')
    
    for r in range(5, 11):
        ws.row_dimensions[r].height = 20
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin_border
            
    # Table of Contents
    ws.cell(row=12, column=1, value="DANH SÁCH CÁC TRANG CẦN ĐẶC TẢ").font = section_font
    ws.merge_cells('A12:G12')
    ws.row_dimensions[12].height = 24
    
    # Headers for Table of Contents
    toc_headers = ["STT", "Mã màn hình", "Tên màn hình đặc tả", "Đường dẫn URL tương ứng", "File Code Component gốc", "Trạng thái", "Ghi chú"]
    for c_idx, h in enumerate(toc_headers, 1):
        cell = ws.cell(row=14, column=c_idx, value=h)
        cell.font = bold_text_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if c_idx in [1, 2, 6] else 'left', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[14].height = 24
    
    # Sample pages
    if not is_admin:
        pages = [
            ("SCR_CLI_001", "Trang chủ", "/vi", "app/[locale]/page.tsx", "Đã gen", "Trang giới thiệu chính, banner, sản phẩm nổi bật, tin tức"),
            ("SCR_CLI_002", "Giới thiệu", "/vi/about", "app/[locale]/about/page.tsx", "Chờ gen", "Giới thiệu lịch sử, sứ mệnh của công ty"),
            ("SCR_CLI_003", "Danh mục Sản phẩm", "/vi/products", "app/[locale]/products/page.tsx", "Chờ gen", "Danh sách sản phẩm kèm bộ lọc, phân trang"),
            ("SCR_CLI_004", "Chi tiết Sản phẩm", "/vi/products/[slug]", "app/[locale]/products/[slug]/page.tsx", "Chờ gen", "Xem chi tiết thông tin sản phẩm, mô tả, ảnh, yêu cầu báo giá"),
            ("SCR_CLI_005", "Liên hệ", "/vi/contact", "app/[locale]/contact/page.tsx", "Chờ gen", "Form gửi thông tin liên hệ và yêu cầu tư vấn"),
            ("SCR_CLI_006", "Khuyến mãi", "/vi/promotions", "app/[locale]/promotions/page.tsx", "Chờ gen", "Danh sách các chương trình khuyến mãi hiện hành"),
            ("SCR_CLI_007", "Hệ thống Showroom", "/vi/showrooms", "app/[locale]/showrooms/page.tsx", "Chờ gen", "Danh sách bản đồ và thông tin liên hệ các showroom")
        ]
    else:
        pages = [
            ("SCR_ADM_001", "Đăng nhập Admin", "/admin/login", "app/admin/login/page.tsx", "Đã gen", "Màn hình đăng nhập hệ thống quản trị"),
            ("SCR_ADM_002", "Bảng điều khiển (Dashboard)", "/admin", "app/admin/page.tsx", "Chờ gen", "Thống kê tổng quan hoạt động website"),
            ("SCR_ADM_003", "Quản lý Sản phẩm", "/admin/products", "app/admin/products/page.tsx", "Chờ gen", "Danh sách sản phẩm, bộ lọc, tìm kiếm, nút CRUD"),
            ("SCR_ADM_004", "Thêm/Sửa Sản phẩm", "/admin/products?new=1", "app/admin/[section]/page.tsx", "Chờ gen", "Form thêm mới hoặc cập nhật thông tin sản phẩm"),
            ("SCR_ADM_005", "Quản lý Danh mục", "/admin/categories", "app/admin/categories/page.tsx", "Chờ gen", "Quản lý cây danh mục đồ gỗ, thiết bị vệ sinh"),
            ("SCR_ADM_006", "Quản lý Thương hiệu", "/admin/brands", "app/admin/brands/page.tsx", "Chờ gen", "Quản lý danh sách đối tác sản xuất, thương hiệu")
        ]
        
    for r_idx, page in enumerate(pages, 15):
        ws.cell(row=r_idx, column=1, value=r_idx - 14).alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=2, value=page[0]).alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=3, value=page[1])
        ws.cell(row=r_idx, column=4, value=page[2])
        ws.cell(row=r_idx, column=5, value=page[3])
        ws.cell(row=r_idx, column=6, value=page[4]).alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=7, value=page[5])
        
        for col in range(1, 8):
            c_cell = ws.cell(row=r_idx, column=col)
            c_cell.font = text_font
            c_cell.border = thin_border
            if page[4] == "Đã gen":
                c_cell.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
            else:
                c_cell.fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        ws.row_dimensions[r_idx].height = 22
        
    # Auto adjust column widths for safety
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def generate_client_template():
    wb = openpyxl.Workbook()
    create_info_intro_sheet(wb, is_admin=False)
    
    # Create SCR_CLI_001_Trang_Chu sheet
    ws = wb.create_sheet(title="SCR_CLI_001_Trang_Chu")
    setup_sheet_style(ws)
    
    r = create_title_section(ws, "Trang chủ", "/vi", "SCR_CLI_001", is_admin=False)
    r = create_image_placeholder_section(ws, r, "pages/desktop/vi-desktop.png")
    r = create_section_header(ws, r, "I. ĐẶC TẢ CHI TIẾT GIAO DIỆN (UI ELEMENTS DESIGN)", is_admin=False)
    
    ui_items = [
        ("Thanh điều hướng (Navbar)", "Navigation Container", "Màu nền: Blur Glassmorphic / White, cố định trên cùng (Sticky Header). Chứa logo, menu liên kết và nút chuyển đổi ngôn ngữ (VI/EN).", "Hiển thị menu điều hướng gồm: Trang chủ, Giới thiệu, Sản phẩm, Khuyến mãi, Showroom, Liên hệ. Hover vào sản phẩm hiển thị dropdown danh mục cấp 2."),
        ("Nút chọn ngôn ngữ", "Dropdown / Button", "Hiển thị 'VI' hoặc 'EN' kèm cờ. Text size: 14px. Padding: 8px 12px.", "Cho phép người dùng chuyển đổi ngôn ngữ website. Khi click, reload trang và thay đổi tiền tố URL tương ứng (ví dụ từ /vi sang /en)."),
        ("Hero Banner Slider", "Carousel Slider", "Chạy ảnh tỷ lệ 21:9, hiển thị full width trên desktop, tự động trượt sau mỗi 5s. Text caption nổi trên ảnh sử dụng Animation fade-in.", "Hiển thị các hình ảnh chiến dịch sản phẩm đồ gỗ phòng khách, nhà tắm. Click nút trên banner sẽ chuyển tới trang sản phẩm tương ứng."),
        ("Danh mục nổi bật (Bento Grid)", "Grid Layout", "Hiển thị 4 danh mục chính (Đồ gỗ nội thất, Thiết bị vệ sinh, Gạch ốp lát, Thiết bị bếp) dạng Bento Grid bất đối xứng. Hover phóng to nhẹ 1.05x.", "Click vào mỗi ô danh mục sẽ chuyển sang trang danh sách sản phẩm và tự động áp dụng bộ lọc cho danh mục đó."),
        ("Form Yêu cầu tư vấn nhanh", "Input Form", "Đặt ở chân trang. Gồm: Họ tên (Input), Số điện thoại (Input, bắt buộc), Email (Input), Ghi chú (Textarea) và nút Gửi.", "Thực hiện validation số điện thoại (đúng định dạng 10 số). Khi click gửi, gọi API POST `/api/quote-requests` lưu thông tin vào CSDL và gửi email thông báo cho Admin.")
    ]
    r = create_ui_spec_table(ws, r, ui_items, is_admin=False)
    
    r = create_section_header(ws, r, "II. CÁC LUỒNG NGHIỆP VỤ & XỬ LÝ (BUSINESS WORKFLOWS)", is_admin=False)
    flows = [
        ("Chuyển đổi ngôn ngữ", "Bước 1: Người dùng click vào nút chọn ngôn ngữ trên Navbar.\nBước 2: Hệ thống hiển thị danh sách các ngôn ngữ hỗ trợ (Tiếng Việt, Tiếng Anh).\nBước 3: Người dùng chọn ngôn ngữ mong muốn.\nBước 4: Hệ thống cập nhật locale trong cookie/URL và tải lại trang với nội dung đã được dịch."),
        ("Gửi yêu cầu tư vấn nhanh", "Bước 1: Người dùng nhập đầy đủ thông tin vào Form tư vấn dưới footer.\nBước 2: Người dùng nhấn nút 'Gửi yêu cầu'.\nBước 3: Frontend kiểm tra dữ liệu đầu vào: Số điện thoại phải bắt buộc và đúng định dạng.\nBước 4: Nếu thông tin hợp lệ, hệ thống gọi API lưu yêu cầu và hiển thị popup 'Gửi yêu cầu thành công'. Nếu lỗi, hiển thị thông báo lỗi tương ứng.")
    ]
    r = create_business_flow_section(ws, r, flows, is_admin=False)
    
    r = create_section_header(ws, r, "III. HƯỚNG DẪN SỬ DỤNG (USER GUIDE)", is_admin=False)
    user_guide_text = (
        "1. Truy cập địa chỉ http://localhost:3000/vi để vào trang chủ.\n"
        "2. Cuộn chuột để xem các khu vực: Banner slide -> Danh mục sản phẩm bento -> Sản phẩm bán chạy -> Tin tức -> Chân trang.\n"
        "3. Rê chuột lên các phần tử menu hoặc danh mục bento để trải nghiệm hiệu ứng mượt mà (hover effect).\n"
        "4. Điền form yêu cầu tư vấn nhanh ở chân trang để nhận liên hệ lại từ phòng kinh doanh trong 24h."
    )
    r = create_user_guide_section(ws, r, user_guide_text)
    
    # Custom column width specifically for detailed sheets
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 45
    
    # Save template sheet for copying
    ws_temp = wb.create_sheet(title="Template_Man_Hinh")
    setup_sheet_style(ws_temp)
    create_title_section(ws_temp, "[TÊN MÀN HÌNH]", "/[route]", "SCR_CLI_XXX", is_admin=False)
    r_t = 11
    r_t = create_image_placeholder_section(ws_temp, r_t, "pages/desktop/placeholder-desktop.png")
    r_t = create_section_header(ws_temp, r_t, "I. ĐẶC TẢ CHI TIẾT GIAO DIỆN (UI ELEMENTS DESIGN)", is_admin=False)
    r_t = create_ui_spec_table(ws_temp, r_t, [
        ("[Tên phần tử]", "[Loại control]", "[Mô tả css, màu sắc, font, validation]", "[Hành động, logic xử lý, API call]")
    ], is_admin=False)
    r_t = create_section_header(ws_temp, r_t, "II. CÁC LUỒNG NGHIỆP VỤ & XỬ LÝ (BUSINESS WORKFLOWS)", is_admin=False)
    r_t = create_business_flow_section(ws_temp, r_t, [
        ("[Tên luồng nghiệp vụ]", "[Các bước thực hiện 1, 2, 3...]")
    ], is_admin=False)
    r_t = create_section_header(ws_temp, r_t, "III. HƯỚNG DẪN SỬ DỤNG (USER GUIDE)", is_admin=False)
    create_user_guide_section(ws_temp, r_t, "1. Hướng dẫn thao tác bước 1\n2. Hướng dẫn thao tác bước 2")
    
    ws_temp.column_dimensions['A'].width = 6
    ws_temp.column_dimensions['B'].width = 25
    ws_temp.column_dimensions['C'].width = 15
    ws_temp.column_dimensions['D'].width = 18
    ws_temp.column_dimensions['E'].width = 25
    ws_temp.column_dimensions['F'].width = 25
    ws_temp.column_dimensions['G'].width = 45

    import os
    os.makedirs("artifacts", exist_ok=True)
    wb.save("artifacts/client_srs_template.xlsx")
    print("Generated artifacts/client_srs_template.xlsx successfully!")

def generate_admin_template():
    wb = openpyxl.Workbook()
    create_info_intro_sheet(wb, is_admin=True)
    
    # Create SCR_ADM_001_Login sheet
    ws = wb.create_sheet(title="SCR_ADM_001_Dang_Nhap")
    setup_sheet_style(ws)
    
    r = create_title_section(ws, "Đăng nhập Admin", "/admin/login", "SCR_ADM_001", is_admin=True)
    r = create_image_placeholder_section(ws, r, "pages/desktop/admin-login-desktop.png")
    r = create_section_header(ws, r, "I. ĐẶC TẢ CHI TIẾT GIAO DIỆN (UI ELEMENTS DESIGN)", is_admin=True)
    
    ui_items = [
        ("Hộp đăng nhập (Login Box)", "Card Container", "Nằm căn giữa màn hình, viền bo tròn 8px, bóng đổ nhẹ (Shadow). Màu nền trắng.", "Chứa form nhập liệu và nút đăng nhập hệ thống quản trị."),
        ("Trường Email", "Input Text", "Placeholder: 'Nhập email của bạn'. Kiểu email validation. Border đỏ khi trống hoặc sai định dạng. Có icon thư ở trước.", "Dùng để nhập email tài khoản quản trị viên. Bắt buộc nhập."),
        ("Trường Mật khẩu", "Input Password", "Placeholder: '••••••••'. Cho phép ẩn/hiện mật khẩu bằng icon mắt ở bên phải. Bắt buộc nhập.", "Dùng để nhập mật khẩu quản trị viên. Bắt buộc nhập."),
        ("Nút Đăng nhập", "Button", "Màu nền: Teal đậm (#0D9488), hover đổi sang Teal sáng (#14B8A6). Text 'Đăng nhập', viết hoa đậm, size 14px.", "Khi người dùng click, kích hoạt kiểm tra dữ liệu đầu vào. Nếu hợp lệ, gửi thông tin email và mật khẩu qua API POST `/api/admin/login` để xác thực.")
    ]
    r = create_ui_spec_table(ws, r, ui_items, is_admin=True)
    
    r = create_section_header(ws, r, "II. CÁC LUỒNG NGHIỆP VỤ & XỬ LÝ (BUSINESS WORKFLOWS)", is_admin=True)
    flows = [
        ("Xác thực đăng nhập", "Bước 1: Người quản trị nhập email và mật khẩu và nhấn 'Đăng nhập'.\nBước 2: Hệ thống kiểm tra tính hợp lệ của email và mật khẩu tại client.\nBước 3: Hệ thống gửi request xác thực tới backend qua API POST `/api/admin/login`.\nBước 4: Backend kiểm tra CSDL Supabase/PostgreSQL, nếu khớp sẽ tạo JWT token / Session cookie và trả về mã thành công.\nBước 5: Frontend nhận kết quả, lưu session và điều hướng người dùng tới trang Dashboard quản trị `/admin`.\nBước 6: Nếu sai tài khoản, hiển thị thông báo lỗi 'Tài khoản hoặc mật khẩu không chính xác'.")
    ]
    r = create_business_flow_section(ws, r, flows, is_admin=True)
    
    r = create_section_header(ws, r, "III. HƯỚNG DẪN SỬ DỤNG (USER GUIDE)", is_admin=True)
    user_guide_text = (
        "1. Mở trình duyệt và truy cập địa chỉ http://localhost:3000/admin/login.\n"
        "2. Nhập chính xác Email và Mật khẩu của tài khoản Admin được cấp.\n"
        "3. Click nút 'Đăng nhập' hoặc nhấn phím Enter trên bàn phím để tiến hành đăng nhập.\n"
        "4. Nếu đăng nhập thành công, bạn sẽ được tự động chuyển hướng vào Trang quản trị chính."
    )
    r = create_user_guide_section(ws, r, user_guide_text)
    
    # Custom column width specifically for detailed sheets
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 45
    
    # Save template sheet for copying
    ws_temp = wb.create_sheet(title="Template_Man_Hinh")
    setup_sheet_style(ws_temp)
    create_title_section(ws_temp, "[TÊN MÀN HÌNH ADMIN]", "/admin/[route]", "SCR_ADM_XXX", is_admin=True)
    r_t = 11
    r_t = create_image_placeholder_section(ws_temp, r_t, "pages/desktop/placeholder-admin-desktop.png")
    r_t = create_section_header(ws_temp, r_t, "I. ĐẶC TẢ CHI TIẾT GIAO DIỆN (UI ELEMENTS DESIGN)", is_admin=True)
    r_t = create_ui_spec_table(ws_temp, r_t, [
        ("[Tên phần tử]", "[Loại control]", "[Mô tả css, màu sắc, font, validation]", "[Hành động, logic xử lý, API call]")
    ], is_admin=True)
    r_t = create_section_header(ws_temp, r_t, "II. CÁC LUỒNG NGHIỆP VỤ & XỬ LÝ (BUSINESS WORKFLOWS)", is_admin=True)
    r_t = create_business_flow_section(ws_temp, r_t, [
        ("[Tên luồng nghiệp vụ]", "[Các bước thực hiện 1, 2, 3...]")
    ], is_admin=True)
    r_t = create_section_header(ws_temp, r_t, "III. HƯỚNG DẪN SỬ DỤNG (USER GUIDE)", is_admin=True)
    create_user_guide_section(ws_temp, r_t, "1. Hướng dẫn thao tác admin bước 1\n2. Hướng dẫn thao tác admin bước 2")
    
    ws_temp.column_dimensions['A'].width = 6
    ws_temp.column_dimensions['B'].width = 25
    ws_temp.column_dimensions['C'].width = 15
    ws_temp.column_dimensions['D'].width = 18
    ws_temp.column_dimensions['E'].width = 25
    ws_temp.column_dimensions['F'].width = 25
    ws_temp.column_dimensions['G'].width = 45

    import os
    os.makedirs("artifacts", exist_ok=True)
    wb.save("artifacts/admin_srs_template.xlsx")
    print("Generated artifacts/admin_srs_template.xlsx successfully!")

if __name__ == "__main__":
    generate_client_template()
    generate_admin_template()
